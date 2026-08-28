import os
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTES_DIR = os.path.join(BASE_DIR, "reportes_generados")
os.makedirs(REPORTES_DIR, exist_ok=True)


def generar_reporte_pdf(titulo: str, datos: list, columnas: list) -> str:
    """
    Genera un reporte PDF a partir de una lista de diccionarios.
    columnas: lista de claves a incluir, en orden.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"reporte_{timestamp}.pdf"
    ruta = os.path.join(REPORTES_DIR, nombre_archivo)

    doc = SimpleDocTemplate(ruta, pagesize=letter)
    estilos = getSampleStyleSheet()
    elementos = [Paragraph(titulo, estilos['Title']), Spacer(1, 20)]

    tabla_datos = [columnas]
    for fila in datos:
        tabla_datos.append([str(fila.get(col, "")) for col in columnas])

    tabla = Table(tabla_datos)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elementos.append(tabla)
    doc.build(elementos)

    return ruta


def generar_reporte_excel(titulo: str, datos: list, columnas: list) -> str:
    """Genera un reporte Excel a partir de una lista de diccionarios."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"reporte_{timestamp}.xlsx"
    ruta = os.path.join(REPORTES_DIR, nombre_archivo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita el nombre de hoja a 31 caracteres

    ws.append(columnas)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for fila in datos:
        ws.append([fila.get(col, "") for col in columnas])

    for columna in ws.columns:
        max_len = max(len(str(c.value)) for c in columna if c.value is not None)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 3

    wb.save(ruta)
    return ruta